import os
import openpyxl
import django
from django.db.models import Q
from django.db import models
from django.core.management.base import BaseCommand

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "Production_Project.settings")
django.setup()
from api.models import Order

orders_dict = {}


def update_orders_dict(new_order):
    if new_order.order_number in orders_dict:
        orders_dict[new_order.order_number].item_type_dict.update(
            new_order.item_type_dict
        )
        orders_dict[new_order.order_number].item_subtype_dict.update(
            new_order.item_subtype_dict
        )
    else:
        orders_dict[new_order.order_number] = new_order
    if Order.objects.filter(order_number=new_order.order_number).exists():
        existing_orders = Order.objects.filter(order_number=new_order.order_number)
        for existing_order in existing_orders:
            if (
                existing_order.archived
                and existing_order.item_type_dict == new_order.item_type_dict
            ):
                new_order.backorder = True
                new_order.save()
        else:
            existing_order.ship_date = new_order.ship_date
            existing_order.item_type_dict = orders_dict[
                new_order.order_number
            ].item_type_dict
            existing_order.item_subtype_dict = orders_dict[
                new_order.order_number
            ].item_subtype_dict
            existing_order.save()
    else:
        orders_dict[new_order.order_number] = new_order
        new_order.save()


def find_workbooks():
    current_directory = os.getcwd()
    files_in_directory = os.listdir(current_directory)
    xlsx_files = [file for file in files_in_directory if file.endswith(".xlsx")]
    if not xlsx_files:
        print("There are no excel workbooks in this directory.")
    else:
        for file in xlsx_files:
            no_files = False
            no_files = check_if_correct_workbook(file)
        if no_files:
            print(
                "There are no QuickBooks Order Reports which are compatible with this program (in this directory)."
            )


def check_if_correct_workbook(file):
    script_directory = os.path.dirname(os.path.abspath(__file__))
    workbook_path = os.path.join(script_directory, file)
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook["Sheet1"]
    cell_value = sheet["B2"].value
    if cell_value != "Assembly":
        no_files = True
        return no_files
    else:
        sort_workbook(sheet)


def sort_workbook(sheet):
    # Initial loop through row[0] until you find a new item (must not include 'Total' in row[0])
    ## Item found: create a secondary loop for the next rows
    ### Check if next_row[0] includes word 'Total
    ##### If yes: break out of the next_row loop, and return to the intial
    ##### If no: check if next_row[1] is an item
    ####### If yes: check if includes word 'Total' in it:
    ########## If yes: continue next_row fn
    ########## If no: update subtype, and continue next_row fn
    ####### If no: create a new item next_row[2, 4, 6, 8] + add it to dict + continue next_row loop
    ########################################################
    ########################################################
    # Initial loop through row[0] until you find a new item (must not include 'Total' in row[0])
    for index, row in enumerate(sheet.iter_rows(min_row=0, min_col=3, max_col=12)):
        item_type = row[0].value
        item_subtype = None
        if item_type is None:
            continue
        elif "Total" in item_type:
            continue
        elif "Ship" in item_type:
            continue
        else:
            ## Item found: create a secondary loop for the next rows
            next_row_number = index + 2
            for next_row in sheet.iter_rows(
                min_row=next_row_number, min_col=3, max_col=12
            ):
                ### Check if next_row[0] includes word 'Total
                if next_row[0].value is not None:
                    if "Total" in next_row[0].value:
                        ##### If yes: break out of the secondary loop, and return to the intial loop
                        break
                else:
                    ##### If no: check if next_row[1] is an item
                    if next_row[1].value is not None:
                        ####### If yes: check if includes word 'Total' in it:
                        if "Total" in next_row[1].value:
                            ########## If yes: continue next_row fn
                            continue
                        else:
                            ########## If no: update subtype, and continue next_row fn
                            if "- Other" in next_row[1].value:
                                item_subtype = next_row[1].value.replace("- Other", "")
                            else:
                                item_subtype = next_row[1].value
                                continue
                    else:
                        ####### If no: create a new item next_row[2, 4, 6, 8] + add it to dict + continue next_row loop
                        item_type_dict = {}
                        item_type_dict[item_type] = next_row[8].value
                        item_subtype_dict = {}
                        if item_subtype is None:
                            item_subtype_dict[item_type] = next_row[8].value
                        else:
                            item_subtype_dict[item_subtype] = next_row[8].value
                        new_order = Order(
                            order_number=next_row[4].value,
                            ship_date=next_row[2].value,
                            customer_name=next_row[6].value,
                            item_type_dict=item_type_dict,
                            item_subtype_dict=item_subtype_dict,
                            packages="",
                            confirmed=False,
                            archived=False,
                        )
                        update_orders_dict(new_order)


find_workbooks()


# def delete_unnecessary_entries():
#     Order.objects.all().delete()


# delete_unnecessary_entries()